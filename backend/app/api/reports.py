from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import io
from ..core.database import get_db
from ..core.security import get_current_user
from ..models.user import User
from ..models.consolidation import ConsolidationRun, Organization

router = APIRouter()

@router.get("/financial-summary")
async def get_financial_summary(organization_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> Dict[str, Any]:
    return {"message": "Financial summary endpoint - implement as needed"}

@router.get("/{run_id}/export/excel")
async def export_to_excel(run_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export consolidation run to Excel with company-by-company detail"""

    run = db.query(ConsolidationRun).join(Organization).filter(
        ConsolidationRun.id == run_id,
        Organization.owner_id == current_user.id
    ).first()

    if not run:
        raise HTTPException(status_code=404, detail="Consolidation run not found")

    # Get actual company-level data (same as board package)
    member_breakdowns = []
    if run.companies_included:
        for company_id in run.companies_included:
            company_data = db.execute(text("SELECT name, goodwill_amount, ownership_percentage FROM companies WHERE id = :company_id"),
                                     {"company_id": company_id}).fetchone()
            if company_data:
                # Query actual revenue
                revenue_query = text("""
                    SELECT COALESCE(SUM(t.credit_amount - t.debit_amount), 0) as revenue
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'REVENUE'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period = :period
                """)
                revenue = db.execute(revenue_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query actual expenses
                expense_query = text("""
                    SELECT COALESCE(SUM(t.debit_amount - t.credit_amount), 0) as expenses
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'EXPENSE'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period = :period
                """)
                expenses = db.execute(expense_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query actual assets
                assets_query = text("""
                    SELECT COALESCE(SUM(t.debit_amount - t.credit_amount), 0) as assets
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'ASSET'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period <= :period
                """)
                assets = db.execute(assets_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query liabilities
                liabilities_query = text("""
                    SELECT COALESCE(SUM(t.credit_amount - t.debit_amount), 0) as liabilities
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'LIABILITY'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period <= :period
                """)
                liabilities = db.execute(liabilities_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Calculate NCI (Non-Controlling Interest)
                ownership_pct = float(company_data[2] or 100.0)
                equity = float(assets - liabilities)
                net_income = float(revenue - expenses)
                nci_percentage = (100.0 - ownership_pct) / 100.0
                nci_equity = equity * nci_percentage
                nci_income = net_income * nci_percentage

                member_breakdowns.append({
                    "company_id": company_id,
                    "company_name": company_data[0],
                    "revenue": float(revenue),
                    "expenses": float(expenses),
                    "net_income": net_income,
                    "assets": float(assets),
                    "liabilities": float(liabilities),
                    "equity": equity,
                    "goodwill_amount": float(company_data[1] or 0),
                    "ownership_percentage": ownership_pct,
                    "nci_equity": nci_equity,
                    "nci_income": nci_income
                })

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border = Border(bottom=Side(style='medium', color='000000'))
    company_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "Financial Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = run.run_name
    ws['A3'] = f"{run.fiscal_year}-{run.fiscal_period:02d}"

    # Calculate NCI totals
    total_nci_equity = sum(m.get('nci_equity', 0) for m in member_breakdowns)
    total_nci_income = sum(m.get('nci_income', 0) for m in member_breakdowns)

    ws['A5'] = "Metric"
    ws['B5'] = "Amount"
    ws['A5'].font = header_font
    ws['B5'].font = header_font
    ws['A5'].fill = header_fill
    ws['B5'].fill = header_fill

    metrics = [
        ("Total Assets", run.total_assets),
        ("Total Liabilities", run.total_liabilities),
        ("Total Equity", run.total_equity),
        ("  Less: Non-Controlling Interest", total_nci_equity),
        ("  Parent Equity", run.total_equity - total_nci_equity),
        ("Total Revenue", run.total_revenue),
        ("Total Expenses", run.total_expenses),
        ("Net Income", run.net_income),
        ("  Less: NCI Income", total_nci_income),
        ("  Parent Net Income", run.net_income - total_nci_income)
    ]

    for i, (label, value) in enumerate(metrics, start=6):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value or 0
        ws[f'B{i}'].number_format = '$#,##0'
        # Indent sub-items
        if label.startswith('  '):
            ws[f'A{i}'].font = Font(italic=True)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # Company-by-Company Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet by Company")
    ws_bs['A1'] = "Balance Sheet - Company Breakdown"
    ws_bs['A1'].font = Font(bold=True, size=14)
    ws_bs['A2'] = f"Period: {run.fiscal_year}-{run.fiscal_period:02d}"

    row = 4
    # Headers
    ws_bs.cell(row=row, column=1, value="Company")
    ws_bs.cell(row=row, column=2, value="Assets")
    ws_bs.cell(row=row, column=3, value="Liabilities")
    ws_bs.cell(row=row, column=4, value="Equity")
    for col in range(1, 5):
        ws_bs.cell(row=row, column=col).font = header_font
        ws_bs.cell(row=row, column=col).fill = header_fill

    # Company data
    row += 1
    total_assets = 0
    total_liabilities = 0
    total_equity = 0

    for member in member_breakdowns:
        ws_bs.cell(row=row, column=1, value=member['company_name'])
        ws_bs.cell(row=row, column=2, value=member['assets']).number_format = '$#,##0'
        ws_bs.cell(row=row, column=3, value=member['liabilities']).number_format = '$#,##0'
        ws_bs.cell(row=row, column=4, value=member['equity']).number_format = '$#,##0'

        total_assets += member['assets']
        total_liabilities += member['liabilities']
        total_equity += member['equity']
        row += 1

    # Totals
    ws_bs.cell(row=row, column=1, value="TOTAL CONSOLIDATED").font = total_font
    ws_bs.cell(row=row, column=2, value=total_assets).number_format = '$#,##0'
    ws_bs.cell(row=row, column=3, value=total_liabilities).number_format = '$#,##0'
    ws_bs.cell(row=row, column=4, value=total_equity).number_format = '$#,##0'
    for col in range(1, 5):
        ws_bs.cell(row=row, column=col).font = total_font
        ws_bs.cell(row=row, column=col).fill = total_fill
        ws_bs.cell(row=row, column=col).border = border

    ws_bs.column_dimensions['A'].width = 30
    ws_bs.column_dimensions['B'].width = 18
    ws_bs.column_dimensions['C'].width = 18
    ws_bs.column_dimensions['D'].width = 18

    # Company-by-Company Income Statement
    ws_is = wb.create_sheet("Income Statement by Company")
    ws_is['A1'] = "Income Statement - Company Breakdown"
    ws_is['A1'].font = Font(bold=True, size=14)
    ws_is['A2'] = f"Period: {run.fiscal_year}-{run.fiscal_period:02d}"

    row = 4
    # Headers
    ws_is.cell(row=row, column=1, value="Company")
    ws_is.cell(row=row, column=2, value="Revenue")
    ws_is.cell(row=row, column=3, value="Expenses")
    ws_is.cell(row=row, column=4, value="Net Income")
    ws_is.cell(row=row, column=5, value="Margin %")
    for col in range(1, 6):
        ws_is.cell(row=row, column=col).font = header_font
        ws_is.cell(row=row, column=col).fill = header_fill

    # Company data
    row += 1
    total_revenue = 0
    total_expenses = 0
    total_net_income = 0

    for member in member_breakdowns:
        margin = (member['net_income'] / member['revenue'] * 100) if member['revenue'] > 0 else 0

        ws_is.cell(row=row, column=1, value=member['company_name'])
        ws_is.cell(row=row, column=2, value=member['revenue']).number_format = '$#,##0'
        ws_is.cell(row=row, column=3, value=member['expenses']).number_format = '$#,##0'
        ws_is.cell(row=row, column=4, value=member['net_income']).number_format = '$#,##0'
        ws_is.cell(row=row, column=5, value=margin).number_format = '0.0"%"'

        # Color code net income
        if member['net_income'] >= 0:
            ws_is.cell(row=row, column=4).font = Font(color='065F46', bold=True)
        else:
            ws_is.cell(row=row, column=4).font = Font(color='DC2626', bold=True)

        total_revenue += member['revenue']
        total_expenses += member['expenses']
        total_net_income += member['net_income']
        row += 1

    # Totals
    consolidated_margin = (total_net_income / total_revenue * 100) if total_revenue > 0 else 0
    ws_is.cell(row=row, column=1, value="TOTAL CONSOLIDATED").font = total_font
    ws_is.cell(row=row, column=2, value=total_revenue).number_format = '$#,##0'
    ws_is.cell(row=row, column=3, value=total_expenses).number_format = '$#,##0'
    ws_is.cell(row=row, column=4, value=total_net_income).number_format = '$#,##0'
    ws_is.cell(row=row, column=5, value=consolidated_margin).number_format = '0.0"%"'
    for col in range(1, 6):
        ws_is.cell(row=row, column=col).font = total_font
        ws_is.cell(row=row, column=col).fill = total_fill
        ws_is.cell(row=row, column=col).border = border

    ws_is.column_dimensions['A'].width = 30
    ws_is.column_dimensions['B'].width = 18
    ws_is.column_dimensions['C'].width = 18
    ws_is.column_dimensions['D'].width = 18
    ws_is.column_dimensions['E'].width = 12

    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"Report_{run.run_name.replace(' ', '_')}_{run.fiscal_year}_{run.fiscal_period:02d}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{run_id}/export/board-package")
async def export_board_package(run_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export comprehensive 12-sheet CFO Board Package"""
    from ..services.excel_export_service import excel_export_service
    from ..models.consolidation import IntercompanyElimination, ConsolidationAdjustment
    from sqlalchemy import text

    run = db.query(ConsolidationRun).join(Organization).filter(
        ConsolidationRun.id == run_id,
        Organization.owner_id == current_user.id
    ).first()

    if not run:
        raise HTTPException(status_code=404, detail="Consolidation run not found")

    # Get member company breakdowns with ACTUAL transaction data
    member_breakdowns = []
    if run.companies_included:
        for company_id in run.companies_included:
            # Get company info
            company_data = db.execute(text("SELECT name, goodwill_amount, ownership_percentage FROM companies WHERE id = :company_id"),
                                     {"company_id": company_id}).fetchone()
            if company_data:
                # Query actual revenue from transactions for this company
                revenue_query = text("""
                    SELECT COALESCE(SUM(t.credit_amount - t.debit_amount), 0) as revenue
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'REVENUE'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period = :period
                """)
                revenue = db.execute(revenue_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query actual expenses
                expense_query = text("""
                    SELECT COALESCE(SUM(t.debit_amount - t.credit_amount), 0) as expenses
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'EXPENSE'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period = :period
                """)
                expenses = db.execute(expense_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query actual assets (debit balances for asset accounts)
                assets_query = text("""
                    SELECT COALESCE(SUM(t.debit_amount - t.credit_amount), 0) as assets
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'ASSET'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period <= :period
                """)
                assets = db.execute(assets_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Query liabilities (credit balances)
                liabilities_query = text("""
                    SELECT COALESCE(SUM(t.credit_amount - t.debit_amount), 0) as liabilities
                    FROM transactions t
                    JOIN company_accounts ca ON t.account_id = ca.id
                    JOIN account_mappings am ON ca.id = am.company_account_id
                    JOIN master_accounts ma ON am.master_account_id = ma.id
                    WHERE t.company_id = :company_id
                    AND ma.account_type = 'LIABILITY'
                    AND t.fiscal_year = :year
                    AND t.fiscal_period <= :period
                """)
                liabilities = db.execute(liabilities_query, {
                    "company_id": company_id,
                    "year": run.fiscal_year,
                    "period": run.fiscal_period
                }).scalar() or 0

                # Calculate NCI (Non-Controlling Interest)
                ownership_pct = float(company_data[2] or 100.0)
                equity = float(assets - liabilities)
                net_income = float(revenue - expenses)
                nci_percentage = (100.0 - ownership_pct) / 100.0
                nci_equity = equity * nci_percentage
                nci_income = net_income * nci_percentage

                member_breakdowns.append({
                    "company_id": company_id,
                    "company_name": company_data[0],
                    "revenue": float(revenue),
                    "expenses": float(expenses),
                    "net_income": net_income,
                    "assets": float(assets),
                    "liabilities": float(liabilities),
                    "equity": equity,
                    "goodwill_amount": float(company_data[1] or 0),
                    "ownership_percentage": ownership_pct,
                    "nci_equity": nci_equity,
                    "nci_income": nci_income
                })

    # Get actual eliminations from database
    eliminations_query = db.execute(
        text("""
            SELECT ie.id, ie.description, ie.elimination_amount, ie.elimination_type,
                   ie.from_company_id, ie.to_company_id, ie.elimination_status,
                   c1.name as from_company_name, c2.name as to_company_name
            FROM intercompany_eliminations ie
            LEFT JOIN companies c1 ON ie.from_company_id = c1.id
            LEFT JOIN companies c2 ON ie.to_company_id = c2.id
            WHERE ie.consolidation_run_id = :run_id
            ORDER BY ie.elimination_amount DESC
        """),
        {"run_id": run_id}
    )
    eliminations = [
        {
            "id": row[0],
            "description": row[1],
            "amount": float(row[2]),
            "type": row[3] or "Intercompany Transaction",
            "from_company_id": row[4],
            "to_company_id": row[5],
            "status": str(row[6]) if row[6] else "eliminated",
            "from_company_name": row[7] or "Unknown",
            "to_company_name": row[8] or "Unknown"
        }
        for row in eliminations_query
    ]

    # Get actual consolidation adjustments from database
    adjustments_query = db.execute(
        text("""
            SELECT ca.id, ca.adjustment_type, ca.description, ca.amount,
                   ca.related_company_id, c.name as company_name
            FROM consolidation_adjustments ca
            LEFT JOIN companies c ON ca.related_company_id = c.id
            WHERE ca.consolidation_run_id = :run_id
            ORDER BY ABS(ca.amount) DESC
        """),
        {"run_id": run_id}
    )
    adjustments = [
        {
            "id": row[0],
            "type": row[1],
            "description": row[2],
            "amount": float(row[3]),
            "related_company_id": row[4],
            "company_name": row[5] or "Consolidated"
        }
        for row in adjustments_query
    ]

    # Get actual account mappings (database-agnostic approach)
    account_mappings = []
    if run.companies_included:
        # Build placeholders for IN clause
        placeholders = ','.join([f':company_id_{i}' for i in range(len(run.companies_included))])
        params = {f'company_id_{i}': company_id for i, company_id in enumerate(run.companies_included)}

        account_mappings_query = db.execute(
            text(f"""
                SELECT am.id, ca.account_number, ca.account_name, ma.account_number as master_account_number,
                       ma.account_name as master_account_name, ma.account_type, c.name as company_name,
                       am.confidence_score
                FROM account_mappings am
                JOIN company_accounts ca ON am.company_account_id = ca.id
                JOIN master_accounts ma ON am.master_account_id = ma.id
                JOIN companies c ON ca.company_id = c.id
                WHERE c.id IN ({placeholders})
                AND am.is_active = true
                ORDER BY c.name, ca.account_number
            """),
            params
        )
        account_mappings = [
            {
                "id": row[0],
                "company_account_number": row[1],
                "company_account_name": row[2],
                "master_account_number": row[3],
                "master_account_name": row[4],
                "account_type": str(row[5]) if row[5] else "",
                "company_name": row[6],
                "confidence_score": float(row[7]) if row[7] else 100.0
            }
            for row in account_mappings_query
        ]

    # Generate Excel file
    excel_file = excel_export_service.generate_board_package(
        run, member_breakdowns, account_mappings, eliminations, adjustments, db
    )

    filename = f"TechCorp_Holdings_Board_Package_{run.fiscal_year}_Q{(run.fiscal_period-1)//3 + 1}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

